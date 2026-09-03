# -*- coding: utf-8 -*-
r"""
web/app.py — FastAPI 入口（MCD AI 内容平台 · v2 全量迁移版）

【项目位置】C:\ideon\mcd-ai-content-platform\web\
【迁移完成度】5/5 页面（00-05）已全部从 Streamlit 迁过来

【启动】
  cd web
  pip install -r requirements.txt
  uvicorn app:app --reload --port 8530
  访问 http://localhost:8530/

【路由清单】
  页面 GET：
    /                00 首页
    /01              01 内容工坊
    /02              02 内容诊断
    /03              03 内容预测
    /04              04 历史洞察
    /05              05 真实结果回流

  API POST（HTMX 表单提交）：
    /api/studio/generate     生成 3 条候选
    /api/studio/select       选择候选（A/B/C）
    /api/studio/ctr-mode     切换 CTR 主流程模式
    /api/studio/l1-toggle    显示/隐藏 L1 实验对比
    /api/diagnosis/diagnose     开始诊断
    /api/diagnosis/rewrite      生成 AI 改写候选
    /api/batch/upload       上传 CSV/Excel
    /api/batch/evaluate     启动批量评估
    /api/batch/download     下载 CSV 结果
    /api/insights/upload       上传历史数据
    /api/feedback/upload       导入回流数据

  GET /health             健康检查

【业务层复用】
  所有 service / repository / prompt 全部走父目录（PROJECT_ROOT），
  不重写业务逻辑，保证单源一致性。
"""

from __future__ import annotations

import io
import os
import sys
import csv
import json
import time
import hmac
import hashlib
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from urllib.parse import quote

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, Response, RedirectResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

# ============================================================
# 路径：让 web/ 能 import 父目录的 services/ / core/ / adapters/ / repositories/
# ============================================================
BASE_DIR = Path(__file__).resolve().parent
TEMPLATES_DIR = BASE_DIR / "templates"
STATIC_DIR = BASE_DIR / "static"

PROJECT_ROOT = BASE_DIR.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

# ============================================================
# 业务层 import（在 sys.path injection 后）
# ============================================================
from core.schemas import (
    TaskInput,
    CHANNELS,
    TARGET_AUDIENCE,
    OBJECTIVES,
    STAGES,
    SCENES,
    TONES,
    ACTIONS,
    PLAN_TYPES,
    COUPON_FLAGS,
)
from core.product_benefit import (
    get_product_categories,
    get_benefit_types,
    get_custom_label,
    options_with_custom,
)
from core.analytics_utils import weighted_ctr

from services.rule_engine import load_rules, check_one, check_candidates
from services.ctr_prediction_service import predict_one, predict_for_candidates
from services.similarity_service import find_similar, summarize_similar
from services.copy_analysis_service import diagnose as svc_diagnose
from services.generation_service import generate, GenerationError, rank_candidates_by_ctr
from services.batch_evaluation_service import (
    parse_batch_file,
    evaluate_batch,
    rows_to_dataframe,
    rows_to_csv_bytes,
    save_predictions_to_records,
)
from services.text_analyzer import (
    add_tokens,
    word_frequency,
    emoji_frequency,
    compare_token,
)
from services.analytics.high_effort_plans import rank_plans
from services.analytics.similarity import find_similar_plans
from services.analytics.daily_trend import daily_aggregate, daily_summary
from services.analytics.owner_compare import owner_compare
from services.feedback_service import (
    import_feedback,
    count as feedback_count,
    aggregate_by_signature,
    read_recent as feedback_read_recent,
)
from services.generation_service import read_recent as gen_read_recent
from services.data_loader import build as data_loader_build

from adapters.ctr_predictor_adapter import predict_l1, predict_l1_status, L1_SUPPORTED_CHANNELS

from prompts import copy_rewrite

try:
    from core.llm_gateway import ProviderRouter
    from ui.llm_status import load_config, is_configured
except Exception:  # noqa: BLE001
    ProviderRouter = None  # type: ignore
    load_config = lambda: {}  # type: ignore
    is_configured = lambda: False  # type: ignore

import web.state as state
from web.state import (
    S_01, S_02, S_03, S_04, S_05,
    store_df, get_df, release_df, reset_01, form_change_signature,
)


# ============================================================
# LLM 状态
# ============================================================
def get_llm_status() -> dict:
    model = os.environ.get("LLM_MODEL", "")
    if not model:
        try:
            cfg = load_config() or {}
            model = cfg.get("model", "") or ""
        except Exception:
            model = ""
    return {
        "configured": is_configured(),
        "model": model or "未配置",
    }


def _build_llm_router():
    """从 ui/llm_status.yaml 构造 ProviderRouter；失败返回 None（走 Demo）。

    容错：
    - yaml 缺失 / api_key 空 → None
    - provider 大小写不匹配（UI 用 "MiniMax"，ProviderRouter 只认 "minimax"）→ 自动 .lower()
    - SDK 未装 / API 异常 → None（不抛错让请求挂掉）
    """
    if ProviderRouter is None:
        return None
    try:
        cfg = load_config() or {}
        if not cfg.get("api_key") or not cfg.get("provider"):
            return None
        return ProviderRouter(
            provider=str(cfg["provider"]).strip().lower(),
            api_key=cfg["api_key"],
            model=cfg.get("model", ""),
        )
    except Exception:
        return None


# ============================================================
# Jinja filters
# ============================================================
def _jinja_safe(v):
    """表格 cell 转字符串：None/NaN → '—'。"""
    if v is None:
        return "—"
    try:
        if isinstance(v, float) and v != v:  # NaN
            return "—"
    except Exception:
        pass
    if isinstance(v, float) and v.is_integer() and abs(v) < 1e15:
        # 不强制转 int；保留浮点精度
        pass
    return v


# ============================================================
# FastAPI + 静态 + 模板
# ============================================================
app = FastAPI(
    title="MCD AI 内容平台",
    description="AI 驱动的内容生产与效果评估平台",
    version="0.3.0",
)

app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

templates = Jinja2Templates(directory=str(TEMPLATES_DIR))
templates.env.filters["safe_cell"] = _jinja_safe


# ============================================================
# 导航配置（单一来源）
# ============================================================
NAV_PAGES = [
    {"id": "home", "route": "/", "name": "首页",
     "subtitle": "", "icon": "home"},
    {"id": "studio", "route": "/studio", "name": "内容工坊",
     "subtitle": "", "icon": "edit"},
    {"id": "diagnosis", "route": "/diagnosis", "name": "内容诊断",
     "subtitle": "", "icon": "stethoscope"},
    {"id": "batch", "route": "/batch", "name": "内容预测",
     "subtitle": "", "icon": "clipboard"},
    {"id": "insights", "route": "/insights", "name": "历史洞察",
     "subtitle": "", "icon": "chart"},
    {"id": "feedback", "route": "/feedback", "name": "结果反哺",
     "subtitle": "", "icon": "refresh"},
    {"id": "settings", "route": "/settings", "name": "字典维护",
     "subtitle": "", "icon": "settings", "hidden_in_nav": True},
]


# ============================================================
# 字典维护鉴权（Phase 40 · 2026-09-02）
# ============================================================
# 简单密码鉴权，无 SSO：
# - 密码：环境变量 MCD_SETTINGS_PASSWORD（默认 ori1026）
# - Cookie：HMAC-SHA256 签名 + 30 天有效，httponly + samesite=lax + path=/settings
# - SECRET：环境变量 MCD_SETTINGS_SECRET（默认占位，生产必改）
# - 仅用于内网个人访问，不要放公网
SETTINGS_PASSWORD = os.environ.get("MCD_SETTINGS_PASSWORD", "ori1026")
SETTINGS_COOKIE_NAME = "mcd_settings_auth"
SETTINGS_COOKIE_MAX_AGE = 30 * 24 * 3600  # 30 天
SETTINGS_SECRET = os.environ.get(
    "MCD_SETTINGS_SECRET", "mcd-default-secret-change-me-in-prod"
)


def _make_settings_cookie() -> str:
    """生成签名 cookie：settings:<ts>:<sig>。"""
    ts = int(time.time())
    payload = f"settings:{ts}"
    sig = hmac.new(
        SETTINGS_SECRET.encode("utf-8"),
        payload.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()[:32]
    return f"{payload}:{sig}"


def _verify_settings_cookie(cookie_value: str) -> bool:
    """验证 cookie：签名匹配 + 在 30 天有效期内。"""
    if not cookie_value:
        return False
    try:
        parts = cookie_value.split(":")
        if len(parts) != 3 or parts[0] != "settings":
            return False
        ts = int(parts[1])
        if (time.time() - ts) > SETTINGS_COOKIE_MAX_AGE:
            return False
        sig = parts[2]
        expected = hmac.new(
            SETTINGS_SECRET.encode("utf-8"),
            f"settings:{ts}".encode("utf-8"),
            hashlib.sha256,
        ).hexdigest()[:32]
        return hmac.compare_digest(sig, expected)
    except Exception:
        return False


def _settings_auth_or_redirect(request: Request):
    """检查 settings 鉴权 cookie。未通过 → 返回 RedirectResponse；否则 None。

    每个 settings 路由开头 `if (r := _settings_auth_or_redirect(request)): return r`
    """
    cookie = request.cookies.get(SETTINGS_COOKIE_NAME, "")
    if not _verify_settings_cookie(cookie):
        next_url = request.url.path
        return RedirectResponse(
            url=f"/settings/login?next={quote(next_url)}",
            status_code=303,
        )
    return None


# ============================================================
# 字典维护（Phase 39 · 2026-09-02）
# ============================================================
# 6 个核心配置：3 个 yaml + 2 个 txt + 1 个 json
# 注意：路径相对 PROJECT_ROOT，不能让用户传（防路径遍历）
DICTIONARIES = [
    {"id": "channel_rules", "name": "渠道规则",
     "path": "config/channel_rules.yaml", "kind": "yaml"},
    {"id": "dimension_weights", "name": "维度权重",
     "path": "config/dimension_weights.yaml", "kind": "yaml"},
    {"id": "coupon_keywords", "name": "含券关键词",
     "path": "config/coupon_keywords.yaml", "kind": "yaml"},
    {"id": "custom_dict", "name": "产品词典",
     "path": "data/custom_dict.txt", "kind": "text"},
    {"id": "stopwords", "name": "停用词",
     "path": "data/stopwords.txt", "kind": "text"},
    {"id": "ctr_baseline", "name": "CTR 基准",
     "path": "data/ctr_baseline.json", "kind": "json"},
]


def _dict_by_id(dict_id: str) -> dict | None:
    for d in DICTIONARIES:
        if d["id"] == dict_id:
            return d
    return None


def base_context(active_id: str) -> dict:
    current = next((p for p in NAV_PAGES if p["id"] == active_id), NAV_PAGES[0])
    llm = get_llm_status()
    return {
        "nav_pages": NAV_PAGES,
        "active_id": active_id,
        "page_title": current["name"],
        "page_subtitle": current["subtitle"],
        "llm_configured": llm["configured"],
        "llm_model": llm["model"],
    }


# ============================================================
# / 00 首页
# ============================================================
@app.get("/", response_class=HTMLResponse)
async def home(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(request, "home.html", base_context("home"))


# ============================================================
# 01 内容工坊
# ============================================================
def _01_context() -> dict:
    """构造 01 页面的完整上下文。"""
    ctx = base_context("studio")

    # options
    product_cats = get_product_categories()
    benefit_types = get_benefit_types()
    custom_label = get_custom_label()
    ctx.update({
        "product_categories": options_with_custom(product_cats),
        "benefit_types": options_with_custom(benefit_types),
        "audience_opts": TARGET_AUDIENCE,
        "objective_opts": OBJECTIVES,
        "stage_opts": STAGES,
        "tone_opts": TONES,
        "channel_opts": CHANNELS,
        "scene_opts": SCENES,
        "action_opts": ACTIONS,
        "plan_type_opts": PLAN_TYPES,
        "coupon_opts": COUPON_FLAGS,
        "ctr_mode_options": S_01["ctr_mode_options"],
        # 状态
        "task_input": S_01["task_input"] or {},
        "candidates": S_01["candidates"],
        "selected_id": S_01["selected_id"],
        "rule_results": S_01["rule_results"],
        "ctr_results": S_01["ctr_results"],
        "similar_summary": S_01["similar_summary"],
        "show_l1": S_01["show_l1"],
        "ctr_mode": S_01["ctr_mode"],
        "last_error": S_01["last_error"],
    })
    # 派生：选中候选 + 其规则/CTR
    sel_id = S_01["selected_id"]
    cand_idx = next(
        (i for i, c in enumerate(S_01["candidates"]) if c.get("id") == sel_id),
        0,
    ) if S_01["candidates"] else 0
    ctx["selected_cand"] = (
        S_01["candidates"][cand_idx] if S_01["candidates"] else {}
    )
    ctx["selected_rule"] = (
        S_01["rule_results"][cand_idx]
        if S_01["rule_results"] and cand_idx < len(S_01["rule_results"])
        else None
    )
    ctx["selected_ctr"] = (
        S_01["ctr_results"][cand_idx]
        if S_01["ctr_results"] and cand_idx < len(S_01["ctr_results"])
        else None
    )
    # 派生 task（用于右侧 channel/stage/objective 模板）
    ctx["task"] = S_01["task_input"] or {}
    # L1 派生（Phase 19 双轨开关）
    l1_ctr = None
    if S_01["show_l1"] and S_01["task_input"] and S_01["candidates"]:
        sel = ctx["selected_cand"]
        if sel:
            try:
                pred, status_l1 = predict_l1(
                    title=sel.get("title", ""),
                    body=sel.get("body", ""),
                    channel=S_01["task_input"].get("channel", "APP Push"),
                    plan_type=S_01["task_input"].get("plan_type", "未知"),
                    coupon=S_01["task_input"].get("coupon", "未知"),
                    workday=S_01["task_input"].get("planned_send_date"),
                )
                l1_ctr = {"pred_ctr": pred, "status": status_l1}
            except Exception:
                l1_ctr = None
    ctx["l1_ctr"] = l1_ctr
    ctx["l1_status_msg"] = (
        f"L1 状态：{predict_l1_status()}；支持渠道：{'、'.join(L1_SUPPORTED_CHANNELS)}"
        if predict_l1_status() == "model"
        else None
    )
    return ctx


@app.get("/studio", response_class=HTMLResponse)
async def page_01(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        request, "pages/01_内容工坊.html", _01_context()
    )


@app.post("/api/studio/generate", response_class=HTMLResponse)
async def api_01_generate(request: Request) -> Response:
    """接收 form，调 generate + check_candidates + predict_for_candidates，返回 /01 页面。"""
    form = await request.form()
    form_dict = {k: form.get(k, "") for k in (
        "product_category", "benefit_type", "audience", "channel",
        "objective", "stage", "scene", "tone", "expected_action",
        "plan_type", "coupon", "planned_send_date", "extra_requirements",
        "text_has_coupon",
    )}
    # planned_send_date 兜底
    if not form_dict.get("planned_send_date"):
        form_dict["planned_send_date"] = None

    try:
        task = TaskInput.from_form(form_dict)
    except ValueError as e:
        S_01["last_error"] = f"必填字段缺失：{e}"
        return RedirectResponse(url="/studio", status_code=303)

    channel_rules, brand_rules = load_rules()
    router = _build_llm_router()
    try:
        candidates = generate(task, router=router, channel_rules=channel_rules)
    except GenerationError as e:
        S_01["last_error"] = f"生成失败：{e}"
        S_01["task_input"] = task.to_dict()
        return RedirectResponse(url="/studio", status_code=303)

    S_01["last_error"] = None
    S_01["task_input"] = task.to_dict()
    S_01["candidates"] = [c.to_dict() for c in candidates]
    rule_results = check_candidates(candidates, task.channel, channel_rules, brand_rules)
    S_01["rule_results"] = [r.to_dict() for r in rule_results]

    ctr_mode = S_01["ctr_mode"]
    if ctr_mode == "l1_model" and predict_l1_status() != "model":
        ctr_mode = "demo"
    # Phase 29 · 2026-09-01 用户翻牌：候选展示固定 A→B→C（不再按 CTR 重排）
    # 反哺影响排序的拍板 #6 改为：CTR 仍展示在右侧"参考结果"，但卡片顺序保持 ABC 固定
    ctr_results = predict_for_candidates(candidates, task, mode=ctr_mode)
    S_01["candidates"] = [c.to_dict() for c in candidates]
    S_01["ctr_results"] = [c.to_dict() if hasattr(c, "to_dict") else c for c in ctr_results]
    # 默认选中 A（Phase 29 用户拍板）
    S_01["selected_id"] = "A"

    first = candidates[0]
    sim_df = find_similar(first.title, first.body, task.channel)
    S_01["similar_summary"] = summarize_similar(sim_df)
    S_01["last_generated_signature"] = form_change_signature(task.to_dict())

    return RedirectResponse(url="/studio", status_code=303)


@app.post("/api/studio/select", response_class=HTMLResponse)
async def api_01_select(request: Request) -> Response:
    form = await request.form()
    sel = form.get("selected_id", "A")
    if sel in ("A", "B", "C"):
        S_01["selected_id"] = sel
    return RedirectResponse(url="/studio", status_code=303)


@app.post("/api/studio/ctr-mode", response_class=HTMLResponse)
async def api_01_ctr_mode(request: Request) -> Response:
    form = await request.form()
    mode = form.get("ctr_mode", "demo")
    if mode in S_01["ctr_mode_options"]:
        if mode == "l1_model" and predict_l1_status() != "model":
            mode = "demo"
        S_01["ctr_mode"] = mode
    return RedirectResponse(url="/studio", status_code=303)


@app.post("/api/studio/l1-toggle", response_class=HTMLResponse)
async def api_01_l1_toggle(request: Request) -> Response:
    form = await request.form()
    S_01["show_l1"] = form.get("show_l1") == "1"
    return RedirectResponse(url="/studio", status_code=303)


# ============================================================
# 02 内容诊断
# ============================================================
def _02_context() -> dict:
    ctx = base_context("diagnosis")
    ctx.update({
        "task": {
            "title": S_02["title"],
            "body": S_02["body"],
            "channel": S_02["channel"],
            "action": S_02["action"],
        },
        "channels": CHANNELS,
        "action_opts": ("regen", "shorten", "direct", "scene", "cta", "safer"),
        "rule": S_02["rule"],
        "diagnose": S_02["diagnose"],
        "ctr": S_02["ctr"],
        "similar_summary": S_02["similar_summary"],
        "similar_rows": S_02["similar_rows"],
        "rewrites": S_02["rewrites"],
        "rewrite_error": S_02["rewrite_error"],
        "rewrite_note": S_02["rewrite_note"],
        "action": S_02["action"],
        "error_msg": S_02["error_msg"],
    })
    return ctx


@app.get("/diagnosis", response_class=HTMLResponse)
async def page_02(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        request, "pages/02_内容诊断.html", _02_context()
    )


def _run_diagnosis():
    title = S_02["title"]
    body = S_02["body"]
    channel = S_02["channel"]

    channel_rules, brand_rules = load_rules()
    rule = check_one(title, body, channel, channel_rules, brand_rules)
    S_02["rule"] = rule.to_dict()

    diag = svc_diagnose(title, body, channel=channel)
    S_02["diagnose"] = diag

    sim_df = find_similar(title, body, channel)
    S_02["similar_summary"] = summarize_similar(sim_df)
    try:
        sim_rows = []
        if sim_df is not None and not sim_df.empty:
            # find_similar_plans 实际列：[plan_id, plan_name, channel, owner, similarity,
            #   n_records, 触达成功, 点击, 加权CTR%]
            # 模板期望 title / body / ctr / similarity，按别名映射 + 容错
            def _cell(row, key):
                if key not in row or pd_is_nan(row[key]):
                    return None
                return row[key]
            for _, row in sim_df.head(5).iterrows():
                sim_rows.append({
                    "title": _cell(row, "plan_name"),
                    "body": None,  # 历史相似未返回正文
                    "ctr": _cell(row, "加权CTR%"),
                    "similarity": _cell(row, "similarity"),
                })
        S_02["similar_rows"] = sim_rows
    except Exception:
        S_02["similar_rows"] = []

    ctr = predict_one(title=title, body=body, channel=channel, mode="demo")
    S_02["ctr"] = ctr.to_dict() if hasattr(ctr, "to_dict") else ctr
    S_02["rewrites"] = []
    S_02["rewrite_error"] = ""


def pd_is_nan(v) -> bool:
    try:
        import math
        if v is None:
            return False
        if isinstance(v, float) and math.isnan(v):
            return True
    except Exception:
        pass
    return False


@app.post("/api/diagnosis/diagnose", response_class=HTMLResponse)
async def api_02_diagnose(request: Request) -> Response:
    form = await request.form()
    S_02["title"] = (form.get("title") or "").strip()
    S_02["body"] = (form.get("body") or "").strip()
    S_02["channel"] = form.get("channel", "APP Push")
    S_02["action"] = form.get("action", "regen")
    S_02["error_msg"] = ""
    if not S_02["body"]:
        S_02["error_msg"] = "正文不能为空"
        return RedirectResponse(url="/diagnosis", status_code=303)
    try:
        _run_diagnosis()
    except Exception as e:  # noqa: BLE001
        S_02["error_msg"] = f"诊断失败：{e}"
    return RedirectResponse(url="/diagnosis", status_code=303)


def _do_rewrite():
    title = S_02["title"]
    body = S_02["body"]
    channel = S_02["channel"]
    action = S_02["action"]
    rewrite_note = S_02["rewrite_note"]

    diag = S_02["diagnose"] or {}
    local = (diag.get("diag") or {}).copy()
    local.setdefault("hit_words", [])
    local.setdefault("miss_top", [])
    local.setdefault("emoji_count", 0)

    router = None
    router = _build_llm_router()

    if router is None or not getattr(router, "api_key", None):
        # Demo 占位（与 Streamlit 版一致：2 条差异化改写）
        S_02["rewrites"] = [
            {"title": f"{channel}改写1",
             "body": f"【改写】{body[:40]}… 立即查看。",
             "reason": f"Demo 占位：{action} 策略示例 1"},
            {"title": f"{channel}改写2",
             "body": f"限时优惠：{body[:30]}… 别错过。",
             "reason": f"Demo 占位：{action} 策略示例 2"},
        ]
        S_02["rewrite_error"] = ""
        return

    channel_rules, _brand_rules = load_rules()
    channel_max = (channel_rules.get("channels") or {}).get(channel, {})
    extra_ctx = f"备注（用户额外要求）：{rewrite_note}" if rewrite_note else ""
    user_prompt = copy_rewrite.build_user_prompt(
        action=action, title=title, body=body,
        channel_max=channel_max, extra_context=extra_ctx,
    )
    full_prompt = f"{copy_rewrite.get_system_prompt(action)}\n\n{user_prompt}"
    try:
        raw = router.call(full_prompt)
        parsed = copy_rewrite.parse_response(raw)
        if "error" in parsed:
            S_02["rewrite_error"] = parsed.get("error", "未知错误")
            S_02["rewrites"] = []
        else:
            S_02["rewrite_error"] = ""
            S_02["rewrites"] = [parsed]
    except Exception as e:  # noqa: BLE001
        S_02["rewrite_error"] = str(e)
        S_02["rewrites"] = []


@app.post("/api/diagnosis/rewrite", response_class=HTMLResponse)
async def api_02_rewrite(request: Request) -> Response:
    if not S_02.get("body"):
        S_02["rewrite_error"] = "请先完成诊断"
        return RedirectResponse(url="/diagnosis", status_code=303)
    form = await request.form()
    S_02["rewrite_note"] = (form.get("rewrite_note") or "").strip()
    try:
        _do_rewrite()
    except Exception as e:  # noqa: BLE001
        S_02["rewrite_error"] = str(e)
    return RedirectResponse(url="/diagnosis", status_code=303)


# ============================================================
# 03 内容预测
# ============================================================
def _03_context() -> dict:
    ctx = base_context("batch")
    rows = []
    for r in S_03["result_rows"]:
        d = dict(r)
        if d.get("ctr_pred") is not None:
            d["ctr_pred_str"] = f"{d['ctr_pred'] * 100:.2f}%"
        else:
            d["ctr_pred_str"] = "—"
        if d.get("ctr_baseline") is not None:
            d["ctr_baseline_str"] = f"{d['ctr_baseline'] * 100:.2f}%"
        else:
            d["ctr_baseline_str"] = "—"
        if d.get("ctr_confidence") is not None:
            d["ctr_confidence_str"] = f"{d['ctr_confidence']:.2f}"
        else:
            d["ctr_confidence_str"] = "—"
        rows.append(d)

    ctx.update({
        "batch": {
            "filename": S_03["filename"],
            "n_rows": S_03["n_rows"],
            "has_title": S_03["has_title"],
            "has_body": S_03["has_body"],
            "has_channel": S_03["has_channel"],
            "n_valid_channels": S_03["n_valid_channels"],
            "n_total_channels": S_03["n_total_channels"],
            "preview_rows": S_03["preview_rows"],
            "can_eval": S_03["can_eval"],
            "save_to_records": S_03["save_to_records"],
            "eval_done": S_03["eval_done"],
            "save_help": "勾选后，评估完成的每行 CTR 预测会落档 records.db（带 signature）。"
                         "下次上传真实 CTR 到 05 时可自动 join 算误差。默认关，按需开启。",
        },
        "result": {
            "n": S_03["n_rows"] if S_03["eval_done"] else 0,
            "n_pass": S_03["n_pass"],
            "n_warn": S_03["n_warn"],
            "n_blocked": S_03["n_blocked"],
            "n_ctr_ok": S_03["n_ctr_ok"],
            "n_err": S_03["n_err"],
            "channel_chips": S_03["channel_chips"],
            "rows": rows,
        },
        "error_msg": S_03["error_msg"],
        "success_msg": S_03["success_msg"],
    })
    return ctx


@app.get("/batch", response_class=HTMLResponse)
async def page_03(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        request, "pages/03_内容预测.html", _03_context()
    )


@app.post("/api/batch/upload", response_class=HTMLResponse)
async def api_03_upload(request: Request, file: UploadFile = File(...)) -> Response:
    S_03["error_msg"] = ""
    S_03["success_msg"] = ""
    try:
        file_bytes = await file.read()
        df = parse_batch_file(file_bytes, file.filename or "")
    except Exception as e:  # noqa: BLE001
        S_03["error_msg"] = f"解析失败：{e}"
        return RedirectResponse(url="/batch", status_code=303)

    S_03["filename"] = file.filename or ""
    # 释放旧 df
    if S_03["df_ref"] is not None:
        release_df(S_03["df_ref"])
    S_03["df_ref"] = store_df(df)
    S_03["n_rows"] = len(df)
    S_03["has_title"] = "title" in df.columns and df["title"].astype(str).str.len().gt(0).any()
    S_03["has_body"] = "body" in df.columns and df["body"].astype(str).str.len().gt(0).any()
    S_03["has_channel"] = "channel" in df.columns

    ch_set = set(df["channel"].dropna().astype(str).tolist()) if S_03["has_channel"] else set()
    valid_ch = ch_set & set(CHANNELS)
    S_03["n_valid_channels"] = len(valid_ch)
    S_03["n_total_channels"] = len(ch_set)

    # 前 5 行预览
    preview = []
    cols = [c for c in ("title", "body", "channel", "plan_type", "coupon") if c in df.columns]
    for _, row in df[cols].head(5).iterrows():
        preview.append({
            "title": str(row.get("title") or ""),
            "body": str(row.get("body") or ""),
            "channel": str(row.get("channel") or ""),
            "plan_type": str(row.get("plan_type") or ""),
            "coupon": str(row.get("coupon") or ""),
        })
    S_03["preview_rows"] = preview
    S_03["can_eval"] = S_03["has_body"] and S_03["has_channel"]
    S_03["eval_done"] = False
    S_03["result_rows"] = []
    S_03["success_msg"] = f"已读取 {len(df)} 行（{file.filename}）"
    return RedirectResponse(url="/batch", status_code=303)


@app.post("/api/batch/evaluate", response_class=HTMLResponse)
async def api_03_evaluate(request: Request) -> Response:
    df = get_df(S_03["df_ref"])
    if df is None:
        S_03["error_msg"] = "请先上传文件"
        return RedirectResponse(url="/batch", status_code=303)

    form = await request.form()
    save_to_records = form.get("save_to_records") == "1"
    S_03["save_to_records"] = save_to_records

    try:
        ctr_mode = "l1_model" if predict_l1_status() == "model" else "demo"
        rows = evaluate_batch(df, ctr_mode=ctr_mode, progress_cb=None)
    except Exception as e:  # noqa: BLE001
        S_03["error_msg"] = f"评估失败：{e}"
        return RedirectResponse(url="/batch", status_code=303)

    S_03["eval_done"] = True
    S_03["result_rows"] = rows
    S_03["result_csv_bytes"] = rows_to_csv_bytes(rows)

    # 聚合统计
    try:
        df_result = rows_to_dataframe(rows)
        n = len(df_result)
        n_blocked = int((df_result["rule_fail_count"] > 0).sum())
        n_warn = int(((df_result["rule_fail_count"] == 0) & (df_result["rule_warn_count"] > 0)).sum())
        n_pass = int(((df_result["rule_fail_count"] == 0) & (df_result["rule_warn_count"] == 0)).sum())
        n_err = int((df_result["error"] != "").sum())
        n_ctr_ok = int(df_result["ctr_pred"].notna().sum())
        S_03["n_pass"] = n_pass
        S_03["n_warn"] = n_warn
        S_03["n_blocked"] = n_blocked
        S_03["n_ctr_ok"] = n_ctr_ok
        S_03["n_err"] = n_err
        if "channel" in df_result.columns:
            ch_counts = df_result["channel"].value_counts().to_dict()
            if ch_counts:
                S_03["channel_chips"] = " · ".join(f"{k} {v}" for k, v in ch_counts.items())
    except Exception:
        S_03["n_pass"] = S_03["n_warn"] = S_03["n_blocked"] = S_03["n_ctr_ok"] = S_03["n_err"] = 0

    S_03["success_msg"] = f"评估完成：{len(rows)} 行"
    if save_to_records:
        try:
            n_saved = save_predictions_to_records(rows)
            if n_saved > 0:
                S_03["success_msg"] += f"；已保存 {n_saved} 条到 records.db"
        except Exception as e:  # noqa: BLE001
            S_03["error_msg"] = f"保存到 records.db 失败：{e}"

    return RedirectResponse(url="/batch", status_code=303)


@app.get("/api/batch/download")
async def api_03_download() -> Response:
    if not S_03.get("result_csv_bytes"):
        raise HTTPException(status_code=404, detail="尚无评估结果")
    return Response(
        content=S_03["result_csv_bytes"],
        media_type="text/csv",
        headers={
            "Content-Disposition": 'attachment; filename="batch_evaluation_result.csv"',
        },
    )


@app.get("/api/batch/template")
async def api_03_template() -> Response:
    """Phase 35 · 2026-09-01 内容预测 Excel 模板下载（含示例行 + 渠道枚举）。"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 未安装，请 pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "内容预测模板"

    # ── 表头（Phase 12 渠道枚举 + parse_batch_file 兼容别名）──
    headers = ["title", "body", "channel", "plan_type", "coupon", "workday_type"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="24292F")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── 示例行（覆盖 3 渠道 + 用券/无券 + 工作日/非工作日）──
    sample_rows = [
        ["夏日新品限时尝鲜，9.9 元起点击立享", "新品上市：板烧鸡腿堡 9.9 元，点击立享优惠。", "APP Push", "常规Plan", "是", "工作日"],
        ["麦当劳会员专属福利", "早安～今日早餐 5 折，点击查看附近门店。", "企微1v1", "AARRPlan", "否", "非工作日"],
        ["106xxxxxxxx", "【麦当劳】新品到店立减 3 元，回复TD退订。", "短信", "常规Plan", "是", "工作日"],
    ]
    for row in sample_rows:
        ws.append(row)

    # ── 列宽自适应 ──
    widths = {"A": 32, "B": 50, "D": 14, "E": 8, "F": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.column_dimensions["C"].width = 22

    # ── 第二 sheet：列说明 ──
    ws_doc = wb.create_sheet("列说明")
    doc_rows = [
        ["列名", "必填", "说明"],
        ["title", "是", "标题文案。兼容别名：标题 / headline / subject"],
        ["body", "是", "正文文案。兼容别名：内容 / content / text"],
        ["channel", "是", f"渠道枚举：{CHANNELS}"],
        ["plan_type", "否", "Plan 类型：AARRPlan / 常规Plan / 未知（缺省走 baseline 兜底）"],
        ["coupon", "否", "实际是否用券：是 / 否 / 未知（未知走兜底）"],
        ["workday_type", "否", "工作日类型：工作日 / 非工作日（缺省按 baseline 兜底）"],
    ]
    for row in doc_rows:
        ws_doc.append(row)
    ws_doc.column_dimensions["A"].width = 16
    ws_doc.column_dimensions["B"].width = 8
    ws_doc.column_dimensions["C"].width = 80
    for cell in ws_doc[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="24292F")

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="batch_prediction_template.xlsx"',
        },
    )


# ============================================================
# 04 历史洞察
# ============================================================
def _safe_int(v, default=0):
    try:
        if v is None:
            return default
        return int(v)
    except (ValueError, TypeError):
        return default


def _plan_detail(df: pd.DataFrame, plan_id: str) -> Optional[dict]:
    """按 Plan ID 精确查询，返回 plan 元数据 + 触达/点击/CTR + 样本标题正文。

    与 rank_plans 一行字段对齐 + 加 title/body 示例，给 Tab 1 「输入 Plan ID 查详情」用。
    """
    if df is None or df.empty or "Plan ID" not in df.columns:
        return None
    sub = df[df["Plan ID"].astype(str) == plan_id]
    if sub.empty:
        return None
    reach = int(sub["触达成功"].sum())
    click = int(sub["点击人次"].sum())
    has_date = "发送日期" in sub.columns
    n_days = sub["发送日期"].dt.date.nunique() if has_date else 0
    sample_title = str(sub["标题"].iloc[0]) if "标题" in sub.columns else ""
    sample_body = str(sub["正文"].iloc[0])[:120] if "正文" in sub.columns else ""
    detail = {
        "plan_id": plan_id,
        "plan_name": str(sub["Plan名称"].iloc[0]) if "Plan名称" in sub.columns else "",
        "channel": str(sub["渠道"].iloc[0]) if "渠道" in sub.columns else "",
        "owner": str(sub["owner"].iloc[0]) if "owner" in sub.columns else "",
        "n_records": int(len(sub)),
        "n_days": int(n_days) if n_days else 0,
        "触达成功": reach,
        "点击": click,
        "加权CTR%": round(weighted_ctr(click, reach), 2),
        "标题字数均值": round(float(sub["标题"].astype(str).str.len().mean()), 1)
        if "标题" in sub.columns else 0,
        "正文字数均值": round(float(sub["正文"].astype(str).str.len().mean()), 1)
        if "正文" in sub.columns else 0,
        "样本标题": sample_title,
        "样本正文": sample_body,
    }
    if "_tokens" in sub.columns:
        tok_set = set()
        for s in sub["_tokens"]:
            tok_set |= set(s)
        detail["覆盖高效词数"] = len(tok_set)
    return detail


def _df_to_rows(df, columns: Optional[list] = None, limit: int = 5000) -> list[dict]:
    """DataFrame → list[dict]，处理 NaN。"""
    if df is None or df.empty:
        return []
    if columns is None:
        columns = list(df.columns)
    safe_cols = [c for c in columns if c in df.columns]
    if limit and len(df) > limit:
        df = df.head(limit)
    rows = []
    for _, row in df[safe_cols].iterrows():
        d = {}
        for c in safe_cols:
            v = row[c]
            try:
                if v is None:
                    d[c] = ""
                    continue
                if isinstance(v, float):
                    import math
                    if math.isnan(v):
                        d[c] = ""
                        continue
                    # 整数浮点化简（仅小数点后为 0 时）
                    if v.is_integer() and abs(v) < 1e15:
                        d[c] = int(v)
                    else:
                        d[c] = round(v, 4)
                else:
                    d[c] = v
            except Exception:
                d[c] = str(v)
        rows.append(d)
    return rows


@app.get("/insights", response_class=HTMLResponse)
async def page_04(request: Request) -> HTMLResponse:
    ctx = base_context("insights")
    df = get_df(S_04["df_ref"])

    # 顶部数据概览（即使没数据也要渲染页面）
    ins = {
        "filename": S_04["filename"],
        "n_rows": S_04["n_rows"],
        "n_has_copy": S_04["n_has_copy"],
        "n_channels": S_04["n_channels"],
        "channels": S_04["channels"],
        "date_range": S_04["date_range"],
    }
    ctx["ins"] = ins
    ctx["error_msg"] = S_04["error_msg"]
    ctx["active_tab"] = request.query_params.get("tab", "rank")
    ctx["params"] = {k: request.query_params.get(k, "") for k in (
        "min_reach", "top_n",
        "wf_min_plans", "wf_top_n", "wf_compare_sel",
        "ef_min_plans", "ef_top_n", "ef_compare_sel",
        "rank_plan_sel",
        "sim_title", "sim_body", "sim_topk",
        "by_channel",
        "oc_min_plans", "oc_min_reach",
    )}
    # 默认值
    ctx["params"].setdefault("min_reach", "1000")
    ctx["params"].setdefault("top_n", "30")
    ctx["params"].setdefault("wf_min_plans", "3")
    ctx["params"].setdefault("wf_top_n", "50")
    ctx["params"].setdefault("wf_compare_sel", "")
    ctx["params"].setdefault("ef_min_plans", "3")
    ctx["params"].setdefault("ef_top_n", "20")
    ctx["params"].setdefault("ef_compare_sel", "")
    ctx["params"].setdefault("rank_plan_sel", "")
    ctx["params"].setdefault("sim_topk", "5")
    ctx["params"].setdefault("oc_min_plans", "3")
    ctx["params"].setdefault("oc_min_reach", "1000")

    if df is None or df.empty:
        return templates.TemplateResponse(
            request, "pages/04_历史洞察.html", ctx
        )

    active_tab = ctx["active_tab"]
    p = ctx["params"]

    # 按 tab 分发计算
    if active_tab == "rank":
        min_reach = _safe_int(p["min_reach"], 1000)
        top_n = _safe_int(p["top_n"], 30)
        out = rank_plans(df, min_reach=min_reach, top_n=top_n)
        if not out.empty:
            show = out.copy()
            show["加权CTR%"] = show["加权CTR%"].apply(lambda v: round(float(v), 2))
        else:
            show = out
        ctx["df_rows"] = _df_to_rows(show, columns=list(show.columns) if not show.empty else None)
        ctx["columns"] = list(show.columns) if not show.empty else []

        # 单 plan 详情（按 Plan ID 精确查询；input → result 一对，仿 wf）
        sel = p["rank_plan_sel"].strip()
        ctx["plan_detail"] = None
        if sel:
            ctx["plan_detail"] = _plan_detail(df, sel)

    elif active_tab == "wf":
        min_plans = _safe_int(p["wf_min_plans"], 3)
        top_n = _safe_int(p["wf_top_n"], 50)
        wf = word_frequency(df, min_plans=min_plans).head(top_n)
        if wf.empty:
            high = low = wf
            ctx["high_rows"] = []
            ctx["low_rows"] = []
            ctx["high_cols"] = ctx["low_cols"] = []
            ctx["wf_words"] = []
        else:
            high = wf[wf["差值"] > 0].head(15)
            low = wf[wf["差值"] < 0].head(15)
            # 数字列保留 4 位
            for sub in (high, low):
                for col in sub.columns:
                    if sub[col].dtype.kind == "f":
                        sub[col] = sub[col].round(4)
            ctx["high_rows"] = _df_to_rows(high)
            ctx["high_cols"] = list(high.columns)
            ctx["low_rows"] = _df_to_rows(low)
            ctx["low_cols"] = list(low.columns)
            ctx["wf_words"] = wf[wf.columns[0]].tolist()[:50]
        ctx["compare"] = None
        sel = p["wf_compare_sel"]
        if sel:
            cmp = compare_token(df, sel)
            if cmp:
                in_block = cmp.get("含", {}) or {}
                out_block = cmp.get("不含", {}) or {}
                ctr_in = float(in_block.get("ctr", 0.0))
                ctr_out = float(out_block.get("ctr", 0.0))
                ctx["compare"] = {
                    "sel_word": sel,
                    "reach_with": int(in_block.get("reach", 0)),
                    "reach_without": int(out_block.get("reach", 0)),
                    "ctr_with": round(ctr_in, 2),
                    "ctr_without": round(ctr_out, 2),
                    "delta_pp": round(ctr_in - ctr_out, 2),
                    "n_plans_with": int(in_block.get("n_plans", 0)),
                    "n_plans_without": int(out_block.get("n_plans", 0)),
                }

    elif active_tab == "ef":
        min_plans = _safe_int(p["ef_min_plans"], 3)
        top_n = _safe_int(p["ef_top_n"], 20)
        ef = emoji_frequency(df, min_plans=min_plans).head(top_n)
        for col in ef.columns:
            if ef[col].dtype.kind == "f":
                ef[col] = ef[col].round(4)
        ctx["df_rows"] = _df_to_rows(ef)
        ctx["columns"] = list(ef.columns)

        # emoji 对比（input → result 一对，仿 wf）；复用 compare_token(col=_emojis)
        sel = p["ef_compare_sel"].strip()
        ctx["ef_compare"] = None
        if sel:
            if "_emojis" not in df.columns:
                df = add_tokens(df)
            cmp = compare_token(df, sel, col="_emojis")
            if cmp:
                in_block = cmp.get("含", {}) or {}
                out_block = cmp.get("不含", {}) or {}
                ctr_in = float(in_block.get("ctr", 0.0))
                ctr_out = float(out_block.get("ctr", 0.0))
                ctx["ef_compare"] = {
                    "sel_emoji": sel,
                    "reach_with": int(in_block.get("reach", 0)),
                    "reach_without": int(out_block.get("reach", 0)),
                    "ctr_with": round(ctr_in, 2),
                    "ctr_without": round(ctr_out, 2),
                    "delta_pp": round(ctr_in - ctr_out, 2),
                    "n_plans_with": int(in_block.get("n_plans", 0)),
                    "n_plans_without": int(out_block.get("n_plans", 0)),
                }

    elif active_tab == "tl":
        if "标题" in df.columns and "Plan ID" in df.columns:
            work = df.copy()
            work["_title_len"] = work["标题"].astype(str).str.len()
            bins = [-1, 0, 5, 10, 15, 20, 1000]
            labels = ["空", "1-5", "6-10", "11-15", "16-20", "21+"]
            work["_bucket"] = pd_cut(work["_title_len"], bins=bins, labels=labels)

            rows = []
            for bucket, sub in work.groupby("_bucket", dropna=False, observed=True):
                reach = int(sub["触达成功"].sum())
                click = int(sub["点击人次"].sum())
                if reach == 0:
                    continue
                n_plans = int(sub["Plan ID"].nunique()) if "Plan ID" in sub.columns else 0
                rows.append({
                    "字数桶": str(bucket),
                    "n_plans": n_plans,
                    "触达成功": reach,
                    "点击": click,
                    "加权CTR%": weighted_ctr(click, reach),
                })
            ctx["df_rows"] = rows
        else:
            ctx["df_rows"] = []

    elif active_tab == "sim":
        q_title = p.get("sim_title", "")
        q_body = p.get("sim_body", "")
        top_k = _safe_int(p["sim_topk"], 5)
        if q_title or q_body:
            sim = find_similar_plans(df, q_title, q_body, top_k=top_k)
            if sim is not None and not sim.empty:
                show_cols = [c for c in ("plan_id", "plan_name", "channel", "ctr", "similarity")
                             if c in sim.columns]
                for col in sim.columns:
                    if sim[col].dtype.kind == "f":
                        sim[col] = sim[col].round(4)
                ctx["df_rows"] = _df_to_rows(sim, columns=show_cols)
                ctx["columns"] = show_cols
            else:
                ctx["df_rows"] = []
                ctx["columns"] = []
        else:
            ctx["df_rows"] = []
            ctx["columns"] = []

    elif active_tab == "daily":
        summary = daily_summary(df) or {}
        ctx["summary"] = summary
        by_channel = bool(p.get("by_channel"))
        out = daily_aggregate(df, channel_col="渠道" if by_channel else None)
        if not out.empty:
            for col in out.columns:
                if out[col].dtype.kind == "f":
                    out[col] = out[col].round(4)
            show_cols = [c for c in (
                "date", "channel", "n_records", "触达成功", "点击", "加权CTR%", "周环比%",
            ) if c in out.columns]
            ctx["df_rows"] = _df_to_rows(out, columns=show_cols)
            ctx["columns"] = show_cols
        else:
            ctx["df_rows"] = []
            ctx["columns"] = []

    elif active_tab == "owner":
        min_plans = _safe_int(p["oc_min_plans"], 3)
        min_reach = _safe_int(p["oc_min_reach"], 1000)
        out = owner_compare(df, min_plans=min_plans, min_reach=min_reach)
        if not out.empty:
            for col in out.columns:
                if out[col].dtype.kind == "f":
                    out[col] = out[col].round(4)
            ctx["df_rows"] = _df_to_rows(out)
            ctx["columns"] = list(out.columns)
        else:
            ctx["df_rows"] = []
            ctx["columns"] = []

    return templates.TemplateResponse(
        request, "pages/04_历史洞察.html", ctx
    )


def pd_cut(series, bins, labels):
    import pandas as _pd
    return _pd.cut(series, bins=bins, labels=labels)


@app.post("/api/insights/upload", response_class=HTMLResponse)
async def api_04_upload(request: Request, file: UploadFile = File(...)) -> Response:
    S_04["error_msg"] = ""
    try:
        file_bytes = await file.read()
        fname = file.filename or ""
        if fname.lower().endswith(".csv"):
            import pandas as pd
            df = pd.read_csv(io.BytesIO(file_bytes))
            meta = {"n_rows": len(df), "sheet_name": "csv", "all_sheets": ["csv"]}
        else:
            df, meta = data_loader_build(file_bytes)
    except Exception as e:  # noqa: BLE001
        S_04["error_msg"] = f"解析失败：{e}"
        return RedirectResponse(url="/insights", status_code=303)

    # 释放旧 df
    if S_04["df_ref"] is not None:
        release_df(S_04["df_ref"])

    try:
        df = add_tokens(df)
    except Exception:
        pass

    S_04["df_ref"] = store_df(df)
    S_04["filename"] = file.filename or ""
    S_04["n_rows"] = meta.get("n_rows", len(df))
    S_04["n_has_copy"] = meta.get("n_has_copy")
    S_04["channels"] = meta.get("channels") or []
    S_04["n_channels"] = len(S_04["channels"])
    if meta.get("date_min"):
        S_04["date_range"] = f"{meta['date_min']} ~ {meta.get('date_max', '')}"
    else:
        S_04["date_range"] = "—"

    return RedirectResponse(url="/insights", status_code=303)


def _parse_xlsx(file_bytes: bytes):
    """Excel 解析（兼容 data_loader.build）。"""
    return data_loader_build(file_bytes)


# ============================================================
# 05 真实结果回流
# ============================================================
def _05_context() -> dict:
    ctx = base_context("feedback")
    ctx.update({
        "error_msg": S_05["error_msg"],
        "success_msg": S_05["success_msg"],
    })

    total = feedback_count()
    summary = {
        "total": total,
        "n_signature": 0,
        "total_reach": 0,
        "overall_ctr": 0.0,
    }
    agg_rows: list[dict] = []
    recent_rows: list[dict] = []
    recent_columns: list[str] = []

    if total > 0:
        agg = aggregate_by_signature()
        summary["n_signature"] = len(agg)
        total_reach = sum(v["reach"] for v in agg.values())
        total_click = sum(v["click"] for v in agg.values())
        summary["total_reach"] = total_reach
        summary["overall_ctr"] = round(total_click / total_reach * 100, 2) if total_reach else 0.0

        rows = []
        for sig, v in agg.items():
            rows.append({
                "signature": sig,
                "channel": v["channel"],
                "n_records": v["n_records"],
                "触达": v["reach"],
                "点击": v["click"],
                "加权CTR%": v["ctr"],
                "订单": v["order_n"],
                "日期范围": f"{v['date_min'] or '—'} ~ {v['date_max'] or '—'}",
            })
        df_agg = pd_DataFrame(rows).sort_values("触达", ascending=False).reset_index(drop=True)
        agg_rows = _df_to_rows(df_agg)

        # 最近 50 条
        recent = feedback_read_recent(limit=50)
        if recent:
            # 去掉 id
            keys = [k for k in recent[0].keys() if k != "id"]
            recent_columns = list(keys)
            for r in recent:
                recent_rows.append({k: r.get(k) for k in keys})

    # join 检查
    join_info = {"n_feedback": 0, "n_joined": 0, "n_only_gen": 0, "n_only_feedback": 0, "error": ""}
    try:
        agg = aggregate_by_signature() if total > 0 else {}
        feedback_sigs = set(agg.keys())
        gen_rows = gen_read_recent(limit=10000)
        gen_sigs = {r.get("signature") for r in gen_rows if r.get("signature")}
        join_info["n_feedback"] = len(feedback_sigs)
        join_info["n_joined"] = len(feedback_sigs & gen_sigs)
        join_info["n_only_gen"] = len(gen_sigs - feedback_sigs)
        join_info["n_only_feedback"] = len(feedback_sigs - gen_sigs)
    except Exception as e:  # noqa: BLE001
        join_info["error"] = f"join 检查失败：{e}"

    ctx.update({
        "summary": summary,
        "agg_rows": agg_rows,
        "recent_rows": recent_rows,
        "recent_columns": recent_columns,
        "join": join_info,
    })
    return ctx


def pd_DataFrame(rows):
    import pandas as _pd
    return _pd.DataFrame(rows)


@app.get("/feedback", response_class=HTMLResponse)
async def page_05(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        request, "pages/05_真实结果回流.html", _05_context()
    )


@app.post("/api/feedback/upload", response_class=HTMLResponse)
async def api_05_upload(
    request: Request,
    file: UploadFile = File(...),
    source_label: str = Form(""),
) -> Response:
    S_05["error_msg"] = ""
    S_05["success_msg"] = ""
    try:
        file_bytes = await file.read()
        result = import_feedback(
            file_bytes,
            filename=file.filename or "",
            source_label=source_label or (file.filename or ""),
        )
    except Exception as e:  # noqa: BLE001
        S_05["error_msg"] = f"导入失败：{e}"
        return RedirectResponse(url="/feedback", status_code=303)

    errs = result.get("errors") or []
    n = result.get("n", 0)
    if errs:
        msg = f"{len(errs)} 条问题（前 5 条）：<br>" + "<br>".join(errs[:5])
        S_05["error_msg"] = msg
    else:
        S_05["success_msg"] = f"已导入 {n} 条回流数据"

    return RedirectResponse(url="/feedback", status_code=303)


# ============================================================
# /settings  字典维护（Phase 39 · 2026-09-02）
# ============================================================
def _read_dict_file(dict_id: str) -> str:
    """读取字典原始文本（保留注释/缩进/yaml 结构）。"""
    d = _dict_by_id(dict_id)
    if not d:
        raise HTTPException(status_code=404, detail=f"未知字典：{dict_id}")
    p = PROJECT_ROOT / d["path"]
    if not p.exists():
        return ""
    try:
        return p.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return p.read_text(encoding="gbk", errors="replace")


def _write_dict_file(dict_id: str, content: str) -> tuple[bool, str]:
    """Atomic write 字典文件（tmp + rename）。ctr_baseline.json 走 JSON 校验。

    返回 (ok, msg)。
    """
    d = _dict_by_id(dict_id)
    if not d:
        return False, f"未知字典：{dict_id}"

    # json 类型先校验
    if d["kind"] == "json":
        try:
            json.loads(content)
        except json.JSONDecodeError as e:
            return False, f"JSON 格式错误：{e}"

    p = PROJECT_ROOT / d["path"]
    p.parent.mkdir(parents=True, exist_ok=True)
    tmp = p.with_suffix(p.suffix + ".tmp")
    try:
        # text 类字典 4 重防御（防止 line ending 累积污染）：
        #   1) CRLF / 孤 CR 全部归一为 LF（兼容任何浏览器 / OS 来源）
        #   2) 空行（用户按回车 + gitattributes normalize 偶发）→ 过滤
        #   3) 每行 rstrip 去行尾空格（保留前导空格用于 indent）
        #   4) 输出 CRLF（gitattributes eol=crlf 一致）
        # 保留 # 注释行（jieba load_userdict 容忍 # 开头）
        # yaml/json 保持原样（保留缩进 / 格式）。
        if d["kind"] == "text":
            raw = content.encode("utf-8").replace(b"\r\n", b"\n").replace(b"\r", b"\n")
            lines = [ln.rstrip() for ln in raw.split(b"\n") if ln.strip()]
            tmp.write_bytes(b"\r\n".join(lines) + b"\r\n")
        else:
            tmp.write_text(content, encoding="utf-8")
        os.replace(tmp, p)
    finally:
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass

    # custom_dict.txt 写入后尝试热重载 jieba（可选，失败不影响保存）
    if dict_id == "custom_dict":
        try:
            from web.services.text_analyzer import load_jieba_dict
            load_jieba_dict()
        except Exception:
            pass
    elif dict_id == "stopwords":
        try:
            from web.services.text_analyzer import load_stopwords
            load_stopwords.cache_clear()
        except Exception:
            pass

    return True, "保存成功"


@app.get("/settings/login", response_class=HTMLResponse)
async def settings_login_page(request: Request, next: str = "/settings", error: str = ""):
    """字典维护登录页（独立模板，简洁居中卡片）。"""
    ctx = base_context("settings")
    ctx["next_url"] = next
    ctx["login_error"] = error
    return templates.TemplateResponse(request, "pages/06_settings_login.html", ctx)


@app.post("/settings/login")
async def settings_login_submit(
    request: Request,
    password: str = Form(...),
    next: str = Form("/settings"),
):
    """校验密码 → 设置 HMAC 签名 cookie → 302 到 next。"""
    if password != SETTINGS_PASSWORD:
        return RedirectResponse(
            url=f"/settings/login?next={quote(next)}&error={quote('密码错误')}",
            status_code=303,
        )
    resp = RedirectResponse(url=next or "/settings", status_code=303)
    resp.set_cookie(
        key=SETTINGS_COOKIE_NAME,
        value=_make_settings_cookie(),
        max_age=SETTINGS_COOKIE_MAX_AGE,
        httponly=True,
        samesite="lax",
        path="/",
    )
    return resp


@app.get("/settings/logout")
async def settings_logout():
    """清 cookie → 302 到 login 页。"""
    resp = RedirectResponse(url="/settings/login", status_code=303)
    resp.delete_cookie(SETTINGS_COOKIE_NAME, path="/")
    return resp


@app.get("/settings", response_class=HTMLResponse)
async def settings_page(request: Request, msg: str = "", err: str = ""):
    """字典维护页面：6 个字典，每个含 textarea + 保存/下载。"""
    if (r := _settings_auth_or_redirect(request)):
        return r
    ctx = base_context("settings")
    dicts = []
    for d in DICTIONARIES:
        item = dict(d)
        try:
            item["content"] = _read_dict_file(d["id"])
            item["size"] = len(item["content"].encode("utf-8"))
            item["lines"] = item["content"].count("\n") + (1 if item["content"] and not item["content"].endswith("\n") else 0)
        except Exception as e:
            item["content"] = ""
            item["size"] = 0
            item["lines"] = 0
            item["error"] = str(e)
        dicts.append(item)
    ctx["dicts"] = dicts
    ctx["flash_msg"] = msg
    ctx["flash_err"] = err
    return templates.TemplateResponse(request, "pages/06_settings.html", ctx)


@app.post("/api/settings/save/{dict_id}")
async def settings_save(request: Request, dict_id: str, content: str = Form(...)):
    if (r := _settings_auth_or_redirect(request)):
        return r
    ok, msg = _write_dict_file(dict_id, content)
    if not ok:
        return RedirectResponse(
            url=f"/settings?err={quote(msg)}",
            status_code=303,
        )
    # 保存成功后自动备份（今天首次保存才备份；失败也不影响保存）
    try:
        from tools.backup_dicts import create_backup_internal
        _, backup_msg = create_backup_internal(days=14)
    except Exception as e:
        backup_msg = f"自动备份失败：{e}"

    d = _dict_by_id(dict_id)
    name = d["name"] if d else dict_id
    full_msg = f"{name} 保存成功 · {backup_msg}"
    return RedirectResponse(
        url=f"/settings?msg={quote(full_msg)}",
        status_code=303,
    )


@app.get("/api/settings/download/{dict_id}")
async def settings_download(request: Request, dict_id: str):
    """直接返回字典文件原文下载。"""
    if (r := _settings_auth_or_redirect(request)):
        return r
    d = _dict_by_id(dict_id)
    if not d:
        raise HTTPException(status_code=404, detail=f"未知字典：{dict_id}")
    p = PROJECT_ROOT / d["path"]
    if not p.exists():
        raise HTTPException(status_code=404, detail=f"文件不存在：{d['path']}")
    return FileResponse(
        path=str(p),
        filename=p.name,
        media_type="application/octet-stream",
    )


# ============================================================
# /api/settings/llm  配置 + 测试连接
# ============================================================
def _mask_api_key(api_key: str) -> str:
    """api_key 脱敏：前 4 + **** + 后 4。短于 8 字符全 ****。"""
    if not api_key:
        return ""
    if len(api_key) < 8:
        return "****"
    return api_key[:4] + "****" + api_key[-4:]


def _write_llm_yaml(cfg: dict) -> None:
    """写 4 字段到家目录下 llm_settings.yaml + 清 lru_cache。

    不写到项目目录是出于两点考虑：
    1) 真实 api_key 不进 git 仓库
    2) 多项目共用一份 LLM 配置
    """
    from ui.llm_status import CONFIG_PATH, _load_yaml
    CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    body = (
        "# LLM 配置（页面右上角 pill 在线配置后保存）\n"
        "# 写到 ~/.mcd-ai/llm_settings.yaml，不进项目目录\n"
        "# 4 字段全部非空启用 LLM 模式，否则走 Demo 占位\n"
        f'provider: "{cfg["provider"]}"\n'
        f'base_url: "{cfg["base_url"]}"\n'
        f'model: "{cfg["model"]}"\n'
        f'api_key: "{cfg["api_key"]}"\n'
    )
    CONFIG_PATH.write_text(body, encoding="utf-8")
    _load_yaml.cache_clear()


# 与 mcd-content-rank/config.py 的 API_PROVIDERS 对齐
# protocol: "openai" / "anthropic" — 决定调哪个 SDK
LLM_PROVIDERS = [
    {"name": "MiniMax",     "base_url": "https://api.minimaxi.com/anthropic",
     "protocol": "anthropic",
     "models": ["MiniMax-M3"]},
    {"name": "火山方舟",     "base_url": "https://ark.cn-beijing.volces.com/api/coding/v3",
     "protocol": "openai",
     "models": ["minimax-m3", "deepseek-v4-flash", "GLM-5.2"]},
    {"name": "百度千帆",     "base_url": "https://qianfan.baidubce.com/v2/coding",
     "protocol": "openai",
     "models": ["qianfan-code-latest"]},
    {"name": "麦当劳AI网关","base_url": "https://ai-gateway-test.mcdchina.net/v1",
     "protocol": "openai",
     "models": ["gemini-3-flash-preview", "gemini-3-pro-image-preview",
                "deepseek-v3", "claude-sonnet-4.6", "claude-haiku-4.5"]},
    {"name": "SiliconFlow", "base_url": "https://api.siliconflow.cn/v1",
     "protocol": "openai",
     "models": ["deepseek-ai/DeepSeek-V3-0324", "Qwen/Qwen2.5-72B-Instruct"]},
    {"name": "OpenAI",      "base_url": "",
     "protocol": "openai",
     "models": ["gpt-4o-mini", "gpt-4o"]},
]


@app.get("/api/settings/llm-modal", response_class=HTMLResponse)
async def api_settings_llm_modal(request: Request):
    """点右上角 pill → 弹 modal。"""
    cfg = load_config() or {}
    return templates.TemplateResponse(
        request,
        "partials/settings_llm_modal.html",
        _modal_context(
            {"provider": cfg.get("provider", ""),
             "base_url": cfg.get("base_url", ""),
             "model": cfg.get("model", ""),
             "api_key": cfg.get("api_key", "")},
            masked_key=_mask_api_key(cfg.get("api_key", "")),
        ),
    )


async def _parse_llm_form(request: Request):
    """提取 + 校验 4 字段。返回 (form_data, errors)。

    model 自动清理 [xxx] / (xxx) 后缀注释，避免 MiniMax-M3[1m] 这种带标注的脏数据。
    """
    import re
    form = await request.form()
    provider = (form.get("provider") or "").strip()
    base_url = (form.get("base_url") or "").strip()
    model = (form.get("model") or "").strip()
    api_key = (form.get("api_key") or "").strip()
    model = re.sub(r"\s*[\[（][^\]）]*[\]）]\s*$", "", model).strip()
    form_data = {"provider": provider, "base_url": base_url,
                  "model": model, "api_key": api_key}
    errors = []
    if not provider: errors.append("provider 必填")
    if not base_url: errors.append("base_url 必填")
    if not model: errors.append("model 必填")
    if not api_key: errors.append("api_key 必填")
    return form_data, errors


def _modal_context(form_data: dict, masked_key: str = "",
                   errors: list = None, test_ok: bool = False, saved: bool = False) -> dict:
    """统一 modal 渲染上下文（避免在多个 endpoint 重复传 providers）。"""
    return {
        "form": form_data,
        "masked_key": masked_key,
        "providers": LLM_PROVIDERS,
        "errors": errors or [],
        "test_ok": test_ok,
        "saved": saved,
    }


def _probe_llm(provider: str, base_url: str, api_key: str, model: str,
               timeout: int = 30):
    """根据 provider protocol 选 SDK 试探连接。返回 (ok, error_msg)。

    protocol=anthropic (MiniMax) 走 anthropic SDK
    protocol=openai    走 openai SDK，base_url 空走 openai 默认
    """
    # 找 protocol
    protocol = "openai"
    for p in LLM_PROVIDERS:
        if p["name"] == provider:
            protocol = p.get("protocol", "openai")
            break

    effective_url = (base_url or "https://api.openai.com/v1").rstrip("/")
    try:
        if protocol == "anthropic":
            import anthropic
            client = anthropic.Anthropic(api_key=api_key, base_url=base_url, timeout=timeout)
            client.messages.create(
                model=model,
                max_tokens=10,
                messages=[{"role": "user", "content": "ping"}],
            )
            return True, ""
        else:
            import openai
            if base_url:
                client = openai.OpenAI(base_url=base_url, api_key=api_key, timeout=timeout)
            else:
                client = openai.OpenAI(api_key=api_key, timeout=timeout)
            client.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": "ping"}],
                max_tokens=10,
            )
            return True, ""
    except Exception as e:
        status = getattr(e, "status_code", None)
        msg = (str(e)[:300] or type(e).__name__)
        detail = f"连接失败: {msg}"
        if status:
            detail += f" [HTTP {status}]"
        endpoint = "/v1/messages" if protocol == "anthropic" else "/chat/completions"
        detail += f"\n请求 URL: {effective_url}{endpoint}"
        if model:
            detail += f"\nModel: {model}"
        detail += f"\nProtocol: {protocol}"
        return False, detail


@app.post("/api/settings/llm/test", response_class=HTMLResponse)
async def api_settings_llm_test(request: Request):
    """只测试连接，不写入文件。"""
    form_data, errors = await _parse_llm_form(request)
    if errors:
        return templates.TemplateResponse(
            request,
            "partials/settings_llm_modal.html",
            _modal_context(form_data,
                           masked_key=_mask_api_key(form_data["api_key"]),
                           errors=errors),
        )
    ok, err = _probe_llm(form_data["provider"], form_data["base_url"],
                         form_data["api_key"], form_data["model"])
    if not ok:
        return templates.TemplateResponse(
            request,
            "partials/settings_llm_modal.html",
            _modal_context(form_data,
                           masked_key=_mask_api_key(form_data["api_key"]),
                           errors=[err]),
        )
    return templates.TemplateResponse(
        request,
        "partials/settings_llm_modal.html",
        _modal_context(form_data,
                       masked_key=_mask_api_key(form_data["api_key"]),
                       test_ok=True),
    )


@app.post("/api/settings/llm", response_class=HTMLResponse)
async def api_settings_llm_save(request: Request):
    """写入 yaml（不重复探测：测试连接已单独跑过，避免 select 改变导致协议错配）。

    写到家目录下，不进项目目录。
    """
    form_data, errors = await _parse_llm_form(request)
    if errors:
        return templates.TemplateResponse(
            request,
            "partials/settings_llm_modal.html",
            _modal_context(form_data,
                           masked_key=_mask_api_key(form_data["api_key"]),
                           errors=errors),
        )

    # 写 yaml + 清缓存（不再探测，避免与 test endpoint 的探测协议不一致）
    _write_llm_yaml(form_data)
    status = get_llm_status()

    # 返回成功 modal（让用户手动关闭）+ 更新 pill 用 hx-swap-oob
    modal_html = templates.TemplateResponse(
        request,
        "partials/settings_llm_modal.html",
        _modal_context(form_data,
                       masked_key=_mask_api_key(form_data["api_key"]),
                       test_ok=True, saved=True),
    ).body.decode("utf-8")
    pill_html = templates.TemplateResponse(
        request,
        "partials/llm_pill.html",
        {"llm_configured": status["configured"],
         "llm_model": status["model"]},
    ).body.decode("utf-8")
    # 用 hx-swap-oob 让 pill 原地更新
    return HTMLResponse(
        f'<div id="settings-modal-slot">{modal_html}</div>'
        f'<div id="llm-pill-slot" hx-swap-oob="outerHTML">{pill_html}</div>'
    )


# ============================================================
# /health
# ============================================================
@app.get("/health")
async def health() -> dict:
    return {"status": "ok", "service": "mcd-ai-content-platform-web"}


# ============================================================
# 启动预热（性能优化 · 2026-09-03）
# ============================================================
# 根因：Jinja2 默认 auto_reload=True，每次渲染都 stat 模板文件；
#       且 01 内容工坊首次渲染要编译 bytecode，导致 /studio 首次 GET 1.33s。
# 修复：服务启动时主动预热 6 个页面模板到 env cache，首次 GET 降至 ~50ms。
# 实测：启动多 ~100ms（一次性），后续每个页面首次访问从 1.3s → ~50ms。
_STARTUP_PAGE_TEMPLATES = (
    "home.html",
    "pages/01_内容工坊.html",
    "pages/02_内容诊断.html",
    "pages/03_内容预测.html",
    "pages/04_历史洞察.html",
    "pages/05_真实结果回流.html",
)


@app.on_event("startup")
async def warmup_page_templates() -> None:
    """服务启动时预热 6 个页面模板，避免首次请求的编译延迟。"""
    for name in _STARTUP_PAGE_TEMPLATES:
        templates.env.get_template(name)


# ============================================================
# 入口
# ============================================================
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app:app", host="127.0.0.1", port=8530, reload=True)
