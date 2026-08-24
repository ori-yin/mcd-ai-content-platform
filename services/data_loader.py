# -*- coding: utf-8 -*-
r"""
services/data_loader.py — Excel 读取 + 消息内容 JSON 解析 + 列映射 + CTR 加权

抽自 C:\ideon\mcd-copy-analyzer\data.py（直接复用，零修改）。
data.py 本身无 Streamlit 依赖，全部为纯函数，可直接搬迁。

口径（与旧项目一致，CLAUDE.md §9）：
- CTR 一律 plan 加权：sum(点击) / sum(触达成功) * 100
- 列名一律"触达成功"（避免和"预计触达"混淆）
"""

from __future__ import annotations

import json
import re
from io import BytesIO
from pathlib import Path
from typing import Tuple

import pandas as pd


_MAIN_SHEET_HINTS = ["sheet", "数据", "明细", "data"]

# 列名模糊映射（兼容 cnn 与旧样本命名），键=标准名
_COL_ALIASES = {
    "发送日期": ["发送日期", "日期", "send", "date"],
    "渠道": ["渠道", "channel"],
    "计划类型": ["计划类型", "plan_type"],
    "Plan ID": ["plan id", "planid", "plan_id"],
    "Plan名称": ["plan名称", "plan_name", "plan 名称"],
    "owner": ["预算owner", "owner", "bu"],
    "是否用券": ["是否用券", "coupon"],
    "预计触达": ["预计触达", "exp_reach"],
    "触达成功": ["触达成功", "reach"],
    "点击人次": ["点击人次", "click"],
    "点击后下单人次": ["点击后下单", "post_click"],
    "订单GC": ["订单gc", "gc"],
    "订单Sales": ["订单sales", "sales"],
    "plan标签": ["消息标题"],  # 内部 plan 命名，非文案
}
_NUM_COLS = ["预计触达", "触达成功", "点击人次", "点击后下单人次", "订单GC", "订单Sales"]


def load_sheet(file_bytes) -> Tuple[pd.DataFrame, str, list]:
    """openpyxl 读主表，保留 emoji。返回 (df_raw, sheet_name, all_sheets)。"""
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        names = list(wb.sheetnames)
        target = next((n for n in names if str(n).strip().lower() in _MAIN_SHEET_HINTS), None)
        if target is None:
            target = max(names, key=lambda n: wb[n].max_row or 0)
        rows = list(wb[target].iter_rows(values_only=True))
    finally:
        wb.close()
    if len(rows) < 2:
        raise ValueError(f"工作表 {target} 没有数据行")
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    return pd.DataFrame(rows[1:], columns=headers), target, names


def map_columns(df: pd.DataFrame) -> pd.DataFrame:
    """列名模糊映射（兼容 cnn 与旧样本命名）。"""
    rename = {}
    existing = set(df.columns)
    for target, keys in _COL_ALIASES.items():
        if target in existing:
            continue
        for c in df.columns:
            cl = str(c).lower().strip().replace(" ", "").replace("_", "")
            if any(str(k).lower().replace(" ", "").replace("_", "") in cl for k in keys):
                rename[c] = target
                break
    return df.rename(columns=rename)


def _extract_from_forms(forms, accept_chain):
    """按 accept_chain（按顺序的多个判断函数）从 forms 里找首个有 value 的项。"""
    if not isinstance(forms, list):
        return None
    for accept in accept_chain:
        for item in forms:
            if accept(str(item.get("code", ""))) and item.get("value"):
                return item["value"]
    return None


_TITLE_CHAIN = (
    lambda c: c == "thing1",
    lambda c: c.startswith("thing"),
    lambda c: not c.startswith("time"),
)
_TEXT_CHAIN = (
    lambda c: c in ("thing5", "short_thing5"),
    lambda c: c.startswith("thing") and c != "thing1",
)


def parse_message(raw) -> Tuple[str, str]:
    """消息内容 JSON -> (标题, 正文)。非 JSON 或空则返回空串。"""
    if raw is None or not isinstance(raw, str) or not raw.strip():
        return "", ""
    try:
        data = json.loads(raw.strip())
    except (json.JSONDecodeError, TypeError):
        return "", ""
    if not isinstance(data, dict):
        return "", ""

    title = data.get("title")
    if not title:
        title = _extract_from_forms(data.get("forms"), _TITLE_CHAIN)
    if not title:
        atts = data.get("attachments")
        if isinstance(atts, list) and atts:
            title = atts[0].get("name", "")

    text = data.get("text")
    if not text:
        text = _extract_from_forms(data.get("forms"), _TEXT_CHAIN)

    # 只有正文没标题：拿正文首句兜底
    if not title and text:
        first = re.split(r"[。！？\n]", str(text).strip())[0].strip()
        title = first if first else str(text)[:20]

    title = str(title).strip() if title else ""
    text = str(text).strip() if text else ""
    for ch in ("\r\n", "\n", "\r"):
        title = title.replace(ch, " ")
        text = text.replace(ch, " ")
    return title, text


def _to_num(s):
    return pd.to_numeric(s, errors="coerce").fillna(0)


def build(file_bytes) -> Tuple[pd.DataFrame, dict]:
    """读取 + 清洗 + 抽文案 + CTR。返回 (df, meta)。

    必备列（映射后）：触达成功 / 点击人次（用于 CTR 计算）。
    """
    df_raw, sheet_name, all_sheets = load_sheet(file_bytes)
    df = map_columns(df_raw.copy())

    # 抽文案：优先 消息内容 JSON；否则回退旧格式的 标题/内容 明文列
    if "消息内容" in df.columns:
        parsed = df["消息内容"].apply(lambda x: pd.Series(parse_message(x), index=["标题", "正文"]))
        df["标题"], df["正文"] = parsed["标题"], parsed["正文"]
    else:
        df["标题"] = df["标题"] if "标题" in df.columns else ""
        df["正文"] = df["内容"] if "内容" in df.columns else ""

    if "发送日期" in df.columns:
        df["发送日期"] = pd.to_datetime(df["发送日期"], errors="coerce")
    for c in _NUM_COLS:
        if c in df.columns:
            df[c] = _to_num(df[c]).astype("int64")

    required = ["触达成功", "点击人次"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"缺少必需列（映射后）：{', '.join(missing)}")

    meta = {
        "sheet_name": sheet_name,
        "all_sheets": all_sheets,
        "n_rows": len(df),
        "n_has_copy": int((df["标题"].str.len() + df["正文"].str.len() > 0).sum()),
        "channels": sorted(df["渠道"].dropna().unique().tolist()) if "渠道" in df.columns else [],
        "owners": sorted(df["owner"].dropna().unique().tolist()) if "owner" in df.columns else [],
        "date_min": None,
        "date_max": None,
    }
    if "发送日期" in df.columns and df["发送日期"].notna().any():
        meta["date_min"] = df["发送日期"].min().date()
        meta["date_max"] = df["发送日期"].max().date()
    return df, meta


__all__ = [
    "load_sheet",
    "map_columns",
    "parse_message",
    "build",
    # 内部 helper（导出供测试用）
    "_extract_from_forms",
    "_to_num",
    "_MAIN_SHEET_HINTS",
    "_COL_ALIASES",
    "_NUM_COLS",
]